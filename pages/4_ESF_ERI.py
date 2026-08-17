import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import datetime


if not st.session_state.get("authentication_status"):
    st.warning("⚠️ Debe iniciar sesión primero.")
    st.page_link("Inicio.py", label="Ir al login", icon="🔐")
    st.stop()

# ── Configuración de página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="Anexos EEFF – ERI y ESF",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .stApp { background-color: #F5F7FA; }
    .header-block {
        background: linear-gradient(135deg, #1E3A5F 0%, #2E6DA4 100%);
        border-radius: 12px; padding: 28px 36px; margin-bottom: 28px; color: white;
    }
    .header-block h1 { font-size: 1.9rem; margin: 0; font-weight: 700; }
    .header-block p  { margin: 6px 0 0; opacity: .85; font-size: .95rem; }
    .metric-card {
        background: white; border-radius: 10px; padding: 20px 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,.08); text-align: center; height: 100%;
    }
    .metric-card .number { font-size: 2.2rem; font-weight: 700; line-height: 1.1; }
    .metric-card .label  { font-size: .82rem; color: #666; margin-top: 4px;
                           text-transform: uppercase; letter-spacing: .05em; }
    .green  { color: #1A9E5C; } .red    { color: #D63B3B; }
    .blue   { color: #2E6DA4; } .orange { color: #E07B20; }
    .purple { color: #7C3AED; }
    .upload-section { background: white; border-radius: 10px; padding: 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,.08); margin-bottom: 20px; }
    .section-title { font-size: 1rem; font-weight: 600; color: #1E3A5F; margin-bottom: 12px;
        padding-bottom: 8px; border-bottom: 2px solid #E8EDF3; }
    .result-block { background: white; border-radius: 10px; padding: 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,.08); margin-top: 20px; }
    #MainMenu, footer, header { visibility: hidden; }
    .block-container { padding-top: 1.5rem; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════
MESES_ORDEN = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
               "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
MESES_ABREV = {"ENERO":"Ene","FEBRERO":"Feb","MARZO":"Mar","ABRIL":"Abr",
               "MAYO":"May","JUNIO":"Jun","JULIO":"Jul","AGOSTO":"Ago",
               "SEPTIEMBRE":"Sep","OCTUBRE":"Oct","NOVIEMBRE":"Nov","DICIEMBRE":"Dic"}

GRUPOS_ERI = ["INGRESOS DE ACTIVIDADES ORDINARIAS","COSTO DE VENTAS","OTROS INGRESOS",
              "GASTOS DE ADMINISTRACION","GASTOS DE VENTA","OTROS GASTOS",
              "INGRESOS FINANCIEROS","GASTOS FINANCIEROS","DIFERENCIA EN CAMBIO NETA",
              "PROVISION DE IMPUESTOS"]
# NOTA: "COSTO DE VENTAS", "GASTOS DE VENTA" y "DIFERENCIA EN CAMBIO NETA" son conceptos
# del catálogo CONCEPTOS_EEFF.xlsx que hoy no existen en ningún archivo de prueba real.
# Si tu script que arma la columna "Grupo" en terceros_ usa un texto distinto para estos
# conceptos, solo hay que ajustar el string aquí (y en NOMBRE_ERI más abajo) para que
# coincida exactamente.

GRUPOS_ESF_ACTIVO = [
    "EFECTIVO Y EQUIVALENTE AL EFECTIVO",
    "OTROS ACTIVOS FINANCIEROS CTE",
    "INVENTARIOS",
    "ACTIVOS POR IMPUESTOS",
    "CUENTAS COMERCIALES POR COBRAR Y OTRAS CUENTAS POR COBRAR CORRIENTES",
    "CUENTAS POR COBRAR A PARTES RELACIONADAS",
    "OTROS ACTIVOS NO FINANCIEROS",
    "OTROS ACTIVOS FINANCIEROS NO CTE",
    "ACTIVOS POR IMPUESTOS DIFERIDOS",
    "ACTIVOS INTANGIBLES DISTINTOS DE LA PLUSVALIA",
    "PROPIEDADES PLANTA Y EQUIPO",
]
GRUPOS_ESF_PASIVO = [
    "OTROS PASIVOS FINANCIEROS",
    "PASIVOS POR IMPUESTOS",
    "CUENTAS COMERCIALES POR PAGAR Y OTRAS CUENTAS POR PAGAR",
    "CUENTAS POR PAGAR A PARTES RELACIONADAS",
    "BENEFICIOS A LOS EMPLEADOS",
    "OTROS PASIVOS NO FINANCIEROS ",
    "CUENTAS COMERCIALES POR PAGAR Y OTRAS CUENTAS POR PAGAR NO CORRIENTES",
    "PASIVOS POR IMPUESTOS DIFERIDOS",
]
GRUPOS_ESF_ORDEN = GRUPOS_ESF_ACTIVO + GRUPOS_ESF_PASIVO
# NOTA: "OTROS ACTIVOS FINANCIEROS CTE", "INVENTARIOS", "OTROS ACTIVOS NO FINANCIEROS",
# "ACTIVOS INTANGIBLES DISTINTOS DE LA PLUSVALIA", "PROPIEDADES PLANTA Y EQUIPO" y
# "OTROS PASIVOS FINANCIEROS" son conceptos del catálogo CONCEPTOS_EEFF.xlsx que hoy no
# existen en ningún archivo de prueba real. Si tu script de clasificación usa otro texto
# exacto para la columna "Grupo", ajusta el string aquí (y en NOMBRE_ESF más abajo).

GRUPOS_LABEL_ESF = {
    "EFECTIVO Y EQUIVALENTE AL EFECTIVO":"🟢 Efectivo y Equiv.",
    "OTROS ACTIVOS FINANCIEROS CTE":"🟢 Otros Activos Fin. Cte",
    "INVENTARIOS":"🟢 Inventarios",
    "CUENTAS COMERCIALES POR COBRAR Y OTRAS CUENTAS POR COBRAR CORRIENTES":"🟢 CxC Corrientes",
    "CUENTAS POR COBRAR A PARTES RELACIONADAS":"🟢 CxC Partes Rel.",
    "ACTIVOS POR IMPUESTOS":"🟢 Activos x Impuestos",
    "OTROS ACTIVOS NO FINANCIEROS":"🟢 Otros Activos No Fin.",
    "ACTIVOS POR IMPUESTOS DIFERIDOS":"🟢 Impuestos Diferidos A.",
    "OTROS ACTIVOS FINANCIEROS NO CTE":"🟢 Otros Activos Fin.",
    "ACTIVOS INTANGIBLES DISTINTOS DE LA PLUSVALIA":"🟢 Intangibles",
    "PROPIEDADES PLANTA Y EQUIPO":"🟢 PP&E",
    "OTROS PASIVOS FINANCIEROS":"🔴 Otros Pasivos Fin.",
    "CUENTAS COMERCIALES POR PAGAR Y OTRAS CUENTAS POR PAGAR":"🔴 CxP Corrientes",
    "CUENTAS COMERCIALES POR PAGAR Y OTRAS CUENTAS POR PAGAR NO CORRIENTES":"🔴 CxP No Corrientes",
    "CUENTAS POR PAGAR A PARTES RELACIONADAS":"🔴 CxP Partes Rel.",
    "PASIVOS POR IMPUESTOS":"🔴 Pasivos x Impuestos",
    "PASIVOS POR IMPUESTOS DIFERIDOS":"🔴 Impuestos Diferidos P.",
    "OTROS PASIVOS NO FINANCIEROS ":"🔴 Otros Pasivos",
    "BENEFICIOS A LOS EMPLEADOS":"🔴 Beneficios Empleados",
}
COLORES_ESF = {
    "EFECTIVO Y EQUIVALENTE AL EFECTIVO":"#2ecc71",
    "OTROS ACTIVOS FINANCIEROS CTE":"#40c977",
    "INVENTARIOS":"#66bb6a",
    "CUENTAS COMERCIALES POR COBRAR Y OTRAS CUENTAS POR COBRAR CORRIENTES":"#27ae60",
    "CUENTAS POR COBRAR A PARTES RELACIONADAS":"#1a9850",
    "ACTIVOS POR IMPUESTOS":"#52b788","ACTIVOS POR IMPUESTOS DIFERIDOS":"#74c69d",
    "OTROS ACTIVOS NO FINANCIEROS":"#86d19a",
    "OTROS ACTIVOS FINANCIEROS NO CTE":"#95d5b2",
    "ACTIVOS INTANGIBLES DISTINTOS DE LA PLUSVALIA":"#b7e4c7",
    "PROPIEDADES PLANTA Y EQUIPO":"#d8f3dc",
    "OTROS PASIVOS FINANCIEROS":"#f1948a",
    "CUENTAS COMERCIALES POR PAGAR Y OTRAS CUENTAS POR PAGAR":"#e74c3c",
    "CUENTAS COMERCIALES POR PAGAR Y OTRAS CUENTAS POR PAGAR NO CORRIENTES":"#c0392b",
    "CUENTAS POR PAGAR A PARTES RELACIONADAS":"#e57373",
    "PASIVOS POR IMPUESTOS":"#ef9a9a","PASIVOS POR IMPUESTOS DIFERIDOS":"#e67e22",
    "OTROS PASIVOS NO FINANCIEROS ":"#d35400","BENEFICIOS A LOS EMPLEADOS":"#9b59b6",
}

# Nombres para hojas formateadas (alineados con el archivo de referencia)
NOMBRE_ESF = {
    "EFECTIVO Y EQUIVALENTE AL EFECTIVO":                                   "  Efectivo y equivalentes al efectivo",
    "OTROS ACTIVOS FINANCIEROS CTE":                                        "  Otros activos financieros",
    "INVENTARIOS":                                                          "  Inventarios",
    "CUENTAS COMERCIALES POR COBRAR Y OTRAS CUENTAS POR COBRAR CORRIENTES": "  Cuentas comerciales por cobrar y otras cuentas por cobrar",
    "CUENTAS POR COBRAR A PARTES RELACIONADAS":                             "  Cuentas por cobrar a partes relacionadas",
    "ACTIVOS POR IMPUESTOS":                                                "  Activos por impuestos",
    "OTROS ACTIVOS NO FINANCIEROS":                                         "  Otros activos no financieros",
    "ACTIVOS POR IMPUESTOS DIFERIDOS":                                      "  Activos por impuestos diferidos",
    "OTROS ACTIVOS FINANCIEROS NO CTE":                                     "  Otros activos financieros",
    "ACTIVOS INTANGIBLES DISTINTOS DE LA PLUSVALIA":                        "  Activos intangibles distintos de la plusvalía",
    "PROPIEDADES PLANTA Y EQUIPO":                                          "  Propiedades, planta y equipo",
    "OTROS PASIVOS FINANCIEROS":                                            "  Otros pasivos financieros",
    "CUENTAS COMERCIALES POR PAGAR Y OTRAS CUENTAS POR PAGAR":              "  Cuentas comerciales por pagar y otras cuentas por pagar",
    "CUENTAS COMERCIALES POR PAGAR Y OTRAS CUENTAS POR PAGAR NO CORRIENTES":"  Cuentas comerciales por pagar y otras cuentas por pagar (NC)",
    "CUENTAS POR PAGAR A PARTES RELACIONADAS":                              "  Cuentas por pagar a partes relacionadas",
    "PASIVOS POR IMPUESTOS":                                                "  Pasivos por impuestos",
    "PASIVOS POR IMPUESTOS DIFERIDOS":                                      "  Pasivos por impuestos diferidos",
    "OTROS PASIVOS NO FINANCIEROS ":                                        "  Otros pasivos no financieros",
    "BENEFICIOS A LOS EMPLEADOS":                                           "  Beneficios a los empleados",
}
NOMBRE_ERI = {
    "INGRESOS DE ACTIVIDADES ORDINARIAS": "Ingresos de actividades ordinarias",
    "COSTO DE VENTAS":                    "Costo de ventas",
    "OTROS INGRESOS":                     "Otros ingresos",
    "GASTOS DE ADMINISTRACION":           "Gastos de administración",
    "GASTOS DE VENTA":                    "Gastos de venta",
    "OTROS GASTOS":                       "Otros gastos",
    "INGRESOS FINANCIEROS":               "Ingresos financieros",
    "GASTOS FINANCIEROS":                 "Gastos financieros",
    "DIFERENCIA EN CAMBIO NETA":          "Diferencia en cambio neta",
    "PROVISION DE IMPUESTOS":             "Ingreso (gasto) por impuesto",
}
GRUPOS_LABEL_ERI = {
    "INGRESOS DE ACTIVIDADES ORDINARIAS":"🟢 Ing. Ordinarios","COSTO DE VENTAS":"🔴 Costo Ventas",
    "OTROS INGRESOS":"🟢 Otros Ingresos","GASTOS DE ADMINISTRACION":"🔴 Gtos. Admón.",
    "GASTOS DE VENTA":"🔴 Gtos. Venta","OTROS GASTOS":"🔴 Otros Gastos",
    "INGRESOS FINANCIEROS":"🔵 Ing. Financieros","GASTOS FINANCIEROS":"🟠 Gtos. Financieros",
    "DIFERENCIA EN CAMBIO NETA":"🟣 Dif. Cambio","PROVISION DE IMPUESTOS":"🟣 Prov. Impuestos",
}
COLORES_ERI = {
    "INGRESOS DE ACTIVIDADES ORDINARIAS":"#2ecc71","COSTO DE VENTAS":"#943126",
    "OTROS INGRESOS":"#27ae60","GASTOS DE ADMINISTRACION":"#e74c3c","GASTOS DE VENTA":"#cb4335",
    "OTROS GASTOS":"#c0392b","INGRESOS FINANCIEROS":"#3498db","GASTOS FINANCIEROS":"#e67e22",
    "DIFERENCIA EN CAMBIO NETA":"#8e44ad","PROVISION DE IMPUESTOS":"#9b59b6","Total general":"#2c3e50",
}

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE DATOS
# ══════════════════════════════════════════════════════════════════════════════
def fmt_cop(val):
    if pd.isna(val): return "-"
    if val < 0: return f"($ {abs(val):,.0f})"
    return f"$ {val:,.0f}"


@st.cache_data(show_spinner="Procesando archivo…")
def procesar_archivo(file_bytes: bytes):
    xls = pd.ExcelFile(BytesIO(file_bytes))
    if "terceros_" not in xls.sheet_names:
        st.error("El archivo no contiene la hoja 'terceros_'.")
        st.stop()

    df = pd.read_excel(BytesIO(file_bytes), sheet_name="terceros_", header=0)
    df.columns = [c.strip() for c in df.columns]
    df["Saldo Mes"] = pd.to_numeric(df["Saldo Mes"], errors="coerce").fillna(0)

    def build_pivot(df_src, grupos):
        df_f = df_src[df_src["Grupo"].isin(grupos)].copy()
        pivot = df_f.groupby(["Grupo","Mes"])["Saldo Mes"].sum().unstack(fill_value=0)
        meses_d = [m for m in MESES_ORDEN if m in pivot.columns]
        pivot = pivot.reindex(columns=meses_d, fill_value=0)
        # Reindex con TODOS los grupos posibles (no solo los que ya tienen datos),
        # para que conceptos sin movimiento en este archivo aparezcan igual con $0.
        pivot = pivot.reindex(grupos, fill_value=0)
        pivot["Total general"] = pivot.sum(axis=1)
        total_row = pivot.sum().to_frame().T
        total_row.index = ["Total general"]
        return pd.concat([pivot, total_row]), df_f

    pivot_eri, df_eri_raw = build_pivot(df, GRUPOS_ERI)

    # ESF
    df_esf_raw = df[df["Grupo"].isin(GRUPOS_ESF_ORDEN)].copy()
    df_esf_raw["Saldo Mes"] = pd.to_numeric(df_esf_raw["Saldo Mes"], errors="coerce").fillna(0)
    pivot_base = df_esf_raw.groupby(["Grupo","Mes"])["Saldo Mes"].sum().unstack(fill_value=0)
    meses_d = [m for m in MESES_ORDEN if m in pivot_base.columns]
    pivot_base = pivot_base.reindex(columns=meses_d, fill_value=0)
    # Igual que en ERI: reindex con TODOS los grupos, no solo los presentes,
    # así los conceptos sin saldo en este archivo quedan en $0 en vez de desaparecer.
    pivot_base = pivot_base.reindex(GRUPOS_ESF_ORDEN, fill_value=0)
    pivot_base["Total general"] = pivot_base.sum(axis=1)

    act_r = [g for g in GRUPOS_ESF_ACTIVO if g in pivot_base.index]
    pas_r = [g for g in GRUPOS_ESF_PASIVO if g in pivot_base.index]
    ta = pivot_base.loc[act_r].sum().to_frame().T; ta.index = ["Total Activo"]
    tp = pivot_base.loc[pas_r].sum().to_frame().T; tp.index = ["Total Pasivo"]
    pivot_esf = pd.concat([pivot_base.loc[act_r], ta, pivot_base.loc[pas_r], tp])

    # Saldo de cierre (último mes) para hoja ESF formateada
    ultimo_mes = meses_d[-1] if meses_d else None
    saldos_esf = {}
    if ultimo_mes:
        for g in GRUPOS_ESF_ORDEN:
            saldos_esf[g] = pivot_base.loc[g, ultimo_mes] if g in pivot_base.index else 0.0

    # Totales ERI acumulados
    totales_eri = {g: pivot_eri.loc[g, "Total general"] if g in pivot_eri.index else 0.0
                   for g in GRUPOS_ERI}

    # Extracción de valores de Patrimonio
    saldos_patrimonio = {
        "capital_emitido": 0.0,
        "superavit_capital": 0.0,
        "utilidad_acumulada": 0.0,
        "reservas": 0.0,
        "utilidad_periodo": 0.0,
        "total_patrimonio": 0.0,
    }

    if ultimo_mes:
        df_m = df[df["Mes"] == ultimo_mes].copy()
        df_m["c_clean"] = df_m["Codigo"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

        # Capital emitido (31 / 3105)
        g_cap = df_m[df_m["Grupo"].astype(str).str.upper().str.contains("CAPITAL EMITIDO|CAPITAL SOCIAL", na=False)]
        if not g_cap.empty:
            cap_emitido = abs(g_cap["Saldo Mes"].sum())
        else:
            m_31 = df_m[df_m["c_clean"].isin(["31", "3105"])]
            if not m_31.empty:
                cap_emitido = abs(m_31["Saldo Mes"].iloc[0])
            else:
                m_31_sub = df_m[df_m["c_clean"].str.startswith("31") & (df_m["c_clean"].str.len() >= 6)]
                cap_emitido = abs(m_31_sub["Saldo Mes"].sum())

        # Superávit de capital (32 / 3205)
        g_sup = df_m[df_m["Grupo"].astype(str).str.upper().str.contains("SUPERAVIT", na=False)]
        if not g_sup.empty:
            superavit_cap = abs(g_sup["Saldo Mes"].sum())
        else:
            m_32 = df_m[df_m["c_clean"].isin(["32", "3205"])]
            if not m_32.empty:
                superavit_cap = abs(m_32["Saldo Mes"].iloc[0])
            else:
                m_32_sub = df_m[df_m["c_clean"].str.startswith("32") & (df_m["c_clean"].str.len() >= 6)]
                superavit_cap = abs(m_32_sub["Saldo Mes"].sum())

        # Utilidad acumulada (36 / 3605 / 37 / 3705)
        g_acum = df_m[df_m["Grupo"].astype(str).str.upper().str.contains("UTILIDAD ACUMULADA|RESULTADOS DE EJERCICIOS ANTERIORES", na=False)]
        if not g_acum.empty:
            util_acum = abs(g_acum["Saldo Mes"].sum())
        else:
            m_36 = df_m[df_m["c_clean"].isin(["37", "3705", "36", "3605"])]
            if not m_36.empty:
                util_acum = abs(m_36["Saldo Mes"].iloc[0])
            else:
                m_36_sub = df_m[df_m["c_clean"].str.startswith(("36", "37")) & (df_m["c_clean"].str.len() >= 6)]
                util_acum = abs(m_36_sub["Saldo Mes"].sum())

        # Reservas (33 / 3305) — concepto del catálogo que hoy no existe en ningún
        # archivo de prueba real; se deja en $0 si no hay cuentas 33xx.
        g_res = df_m[df_m["Grupo"].astype(str).str.upper().str.contains("RESERVAS", na=False)]
        if not g_res.empty:
            reservas = abs(g_res["Saldo Mes"].sum())
        else:
            m_33 = df_m[df_m["c_clean"].isin(["33", "3305"])]
            if not m_33.empty:
                reservas = abs(m_33["Saldo Mes"].iloc[0])
            else:
                m_33_sub = df_m[df_m["c_clean"].str.startswith("33") & (df_m["c_clean"].str.len() >= 6)]
                reservas = abs(m_33_sub["Saldo Mes"].sum())

        # Utilidad del periodo (desde el ERI) — incluye costo de ventas, gastos de
        # venta y diferencia en cambio neta.
        ing_ord    = abs(totales_eri.get("INGRESOS DE ACTIVIDADES ORDINARIAS", 0))
        costo_vtas = abs(totales_eri.get("COSTO DE VENTAS", 0))
        otros_ing  = abs(totales_eri.get("OTROS INGRESOS", 0))
        gtos_adm   = abs(totales_eri.get("GASTOS DE ADMINISTRACION", 0))
        gtos_venta = abs(totales_eri.get("GASTOS DE VENTA", 0))
        otros_gto  = abs(totales_eri.get("OTROS GASTOS", 0))
        ing_fin    = abs(totales_eri.get("INGRESOS FINANCIEROS", 0))
        gto_fin    = abs(totales_eri.get("GASTOS FINANCIEROS", 0))
        # Diferencia en cambio neta: puede ser ganancia o pérdida, se usa el signo
        # natural del saldo (crédito = ganancia, débito = pérdida).
        dif_cambio = -totales_eri.get("DIFERENCIA EN CAMBIO NETA", 0)
        provision  = abs(totales_eri.get("PROVISION DE IMPUESTOS", 0))
        ganancia_bruta = ing_ord - costo_vtas
        util_ai   = (ganancia_bruta + otros_ing - gtos_adm - gtos_venta - otros_gto
                     + ing_fin - gto_fin + dif_cambio)
        util_per  = util_ai - provision

        total_patrimonio = cap_emitido + superavit_cap + util_acum + util_per + reservas

        saldos_patrimonio = {
            "capital_emitido": cap_emitido,
            "superavit_capital": superavit_cap,
            "utilidad_acumulada": util_acum,
            "utilidad_periodo": util_per,
            "reservas": reservas,
            "total_patrimonio": total_patrimonio,
        }

    return df_eri_raw, pivot_eri, df_esf_raw, pivot_esf, saldos_esf, saldos_patrimonio, totales_eri, ultimo_mes


# ══════════════════════════════════════════════════════════════════════════════
# GENERACIÓN EXCEL FORMATEADO
# ══════════════════════════════════════════════════════════════════════════════
FMT_COP_XL  = '_-"$" * #,##0_-;\\-"$" * #,##0_-;_-"$" * "-"_-;_-@_-'
FMT_SUB_XL  = '#,##0_);\\(#,##0\\);"-       "'

def F(bold=False, size=13, name="Calibri"):
    return Font(bold=bold, size=size, name=name)

def A(h="general", v="center"):
    return Alignment(horizontal=h, vertical=v)

def thin():   return Side(border_style="thin")
def double(): return Side(border_style="double")

def set_val(ws, cell_ref, value, bold=False, size=13, fmt=None,
            halign="general", border_bottom=None):
    c = ws[cell_ref]
    c.value = value if value != 0 else (value if value is not None else None)
    c.font  = F(bold=bold, size=size)
    if halign != "general": c.alignment = A(halign)
    if fmt: c.number_format = fmt
    if border_bottom == "thin":   c.border = Border(bottom=thin())
    if border_bottom == "double": c.border = Border(bottom=double())


def generar_hoja_esf(ws, empresa, nit, periodo, saldos, saldos_patrimonio=None):
    if saldos_patrimonio is None:
        saldos_patrimonio = {}

    anchos = {"A":4,"B":49.6,"C":6.1,"D":17.7,"E":2.6,"F":17.7,
              "G":2.6,"H":12.1,"I":49.6,"J":6.1,"K":17.7,"L":2.6,"M":17.7,"N":12}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w
    for r in range(1, 42):
        ws.row_dimensions[r].height = 15.0
    ws.row_dimensions[1].height = 26.1
    for r in [2,3,4,5]: ws.row_dimensions[r].height = 20.1

    # Encabezado
    for r, txt, bold in [(1,empresa,True),(2,nit,False),
                         (3,"ESTADO DE SITUACIÓN FINANCIERA",True),
                         (4,periodo,False),(5,"(En pesos colombianos - $)",False)]:
        c = ws[f"B{r}"]; c.value = txt; c.font = F(bold=bold, size=18)

    # Cabecera columnas fila 7-8
    for col, val in [("C","NOTA"),("D",periodo),("F","Período anterior"),
                     ("J","NOTA"),("K",periodo),("M","Período anterior")]:
        ws[f"{col}7"].value = val
        ws[f"{col}7"].font  = F(bold=True, size=13)
        ws[f"{col}7"].alignment = A("center")
    for col in ["D","F","K","M"]:
        ws[f"{col}8"].value = "$"; ws[f"{col}8"].font = F(bold=True,size=13)
        ws[f"{col}8"].alignment = A("center")

    # Títulos sección
    ws["B9"].value = "ACTIVOS";              ws["B9"].font = F(bold=True, size=14)
    ws["I9"].value = "PASIVOS Y PATRIMONIO"; ws["I9"].font = F(bold=True, size=14)
    ws["B11"].value = "Activos corrientes:"; ws["B11"].font = F(size=13)
    ws["I11"].value = "Pasivos corrientes:"; ws["I11"].font = F(size=13)

    # Siempre devuelve el valor (aunque sea 0). El formato contable FMT_COP_XL
    # ya se encarga de mostrar "-" cuando el saldo es cero, para que el concepto
    # aparezca en la plantilla en vez de quedar oculto.
    def v(g): return abs(saldos.get(g, 0))

    # ── Activos corrientes (col B/D) ──────────────────────────────────────────
    act_cte_map = [
        (12, "EFECTIVO Y EQUIVALENTE AL EFECTIVO"),
        (13, "OTROS ACTIVOS FINANCIEROS CTE"),
        (14, "ACTIVOS POR IMPUESTOS"),
        (15, "CUENTAS COMERCIALES POR COBRAR Y OTRAS CUENTAS POR COBRAR CORRIENTES"),
        (16, "CUENTAS POR COBRAR A PARTES RELACIONADAS"),
        (17, "INVENTARIOS"),
        (18, "OTROS ACTIVOS NO FINANCIEROS"),
    ]
    for row_n, g in act_cte_map:
        ws[f"B{row_n}"].value = NOMBRE_ESF[g]; ws[f"B{row_n}"].font = F(size=13)
        ws[f"B{row_n}"].alignment = A("left")
        ws[f"D{row_n}"].value = v(g); ws[f"D{row_n}"].number_format = FMT_COP_XL
        ws[f"D{row_n}"].font = F(size=13)

    total_act_cte = sum(abs(saldos.get(g,0)) for _,g in act_cte_map)
    ws["B19"].value = "Total activos corrientes"; ws["B19"].font = F(size=13)
    ws["D19"].value = total_act_cte; ws["D19"].number_format = FMT_SUB_XL
    ws["D19"].font = F(size=13); ws["D19"].border = Border(bottom=thin())

    # ── Activos no corrientes (col B/D) ──────────────────────────────────────
    ws["B24"].value = "Activos no corrientes:"; ws["B24"].font = F(size=13)
    act_nct_map = [
        (25, "OTROS ACTIVOS FINANCIEROS NO CTE"),
        (26, "ACTIVOS POR IMPUESTOS DIFERIDOS"),
        (27, "ACTIVOS INTANGIBLES DISTINTOS DE LA PLUSVALIA"),
        (28, "PROPIEDADES PLANTA Y EQUIPO"),
    ]
    for row_n, g in act_nct_map:
        ws[f"B{row_n}"].value = NOMBRE_ESF[g]; ws[f"B{row_n}"].font = F(size=13)
        ws[f"B{row_n}"].alignment = A("left")
        ws[f"D{row_n}"].value = v(g); ws[f"D{row_n}"].number_format = FMT_COP_XL
        ws[f"D{row_n}"].font = F(size=13)

    total_act_nct = sum(abs(saldos.get(g,0)) for _,g in act_nct_map)
    ws["B29"].value = "Total activos no corrientes"; ws["B29"].font = F(size=13)
    ws["D29"].value = total_act_nct; ws["D29"].number_format = FMT_SUB_XL
    ws["D29"].font = F(size=13); ws["D29"].border = Border(bottom=thin())

    total_act = total_act_cte + total_act_nct
    ws["B38"].value = "TOTAL ACTIVOS"; ws["B38"].font = F(bold=True, size=14)
    ws["D38"].value = total_act; ws["D38"].number_format = FMT_COP_XL
    ws["D38"].font = F(bold=True, size=14); ws["D38"].border = Border(bottom=double())

    # ── Pasivos corrientes (col I/K) ──────────────────────────────────────────
    pas_cte_map = [
        (12, "OTROS PASIVOS NO FINANCIEROS "),
        (13, "OTROS PASIVOS FINANCIEROS"),
        (14, "PASIVOS POR IMPUESTOS"),
        (15, "CUENTAS COMERCIALES POR PAGAR Y OTRAS CUENTAS POR PAGAR"),
        (16, "CUENTAS POR PAGAR A PARTES RELACIONADAS"),
        (18, "BENEFICIOS A LOS EMPLEADOS"),
    ]
    for row_n, g in pas_cte_map:
        ws[f"I{row_n}"].value = NOMBRE_ESF[g]; ws[f"I{row_n}"].font = F(size=13)
        ws[f"I{row_n}"].alignment = A("left")
        ws[f"K{row_n}"].value = v(g); ws[f"K{row_n}"].number_format = FMT_COP_XL
        ws[f"K{row_n}"].font = F(size=13)

    total_pas_cte = sum(abs(saldos.get(g,0)) for _,g in pas_cte_map)
    ws["I20"].value = "Total pasivos corrientes"; ws["I20"].font = F(size=13)
    ws["K20"].value = total_pas_cte; ws["K20"].number_format = FMT_SUB_XL
    ws["K20"].font = F(size=13); ws["K20"].border = Border(bottom=thin())

    # ── Pasivos no corrientes (col I/K) ──────────────────────────────────────
    ws["I22"].value = "Pasivos no corrientes:"; ws["I22"].font = F(size=13)
    pas_nct_map = [
        (23, "CUENTAS COMERCIALES POR PAGAR Y OTRAS CUENTAS POR PAGAR NO CORRIENTES"),
        (24, "PASIVOS POR IMPUESTOS DIFERIDOS"),
    ]
    for row_n, g in pas_nct_map:
        ws[f"I{row_n}"].value = NOMBRE_ESF[g]; ws[f"I{row_n}"].font = F(size=13)
        ws[f"I{row_n}"].alignment = A("left")
        ws[f"K{row_n}"].value = v(g); ws[f"K{row_n}"].number_format = FMT_COP_XL
        ws[f"K{row_n}"].font = F(size=13)

    total_pas_nct = sum(abs(saldos.get(g,0)) for _,g in pas_nct_map)
    ws["I26"].value = "Total pasivos no corrientes"; ws["I26"].font = F(size=13)
    ws["K26"].value = total_pas_nct; ws["K26"].number_format = FMT_SUB_XL
    ws["K26"].font = F(size=13); ws["K26"].border = Border(bottom=thin())

    total_pas = total_pas_cte + total_pas_nct
    ws["I28"].value = "TOTAL PASIVOS"; ws["I28"].font = F(bold=True, size=14)
    ws["K28"].value = total_pas; ws["K28"].number_format = FMT_COP_XL
    ws["K28"].font = F(bold=True, size=14); ws["K28"].border = Border(bottom=double())

    # ── Patrimonio (col I/K) ──────────────────────────────────────────────────
    ws["I30"].value = "Patrimonio:"; ws["I30"].font = F(size=13)
    cap_emitido  = saldos_patrimonio.get("capital_emitido", 0.0)
    superavit    = saldos_patrimonio.get("superavit_capital", 0.0)
    util_acum    = saldos_patrimonio.get("utilidad_acumulada", 0.0)
    util_periodo = saldos_patrimonio.get("utilidad_periodo", 0.0)
    reservas     = saldos_patrimonio.get("reservas", 0.0)
    total_patrimonio = saldos_patrimonio.get(
        "total_patrimonio", cap_emitido + superavit + util_acum + util_periodo + reservas)

    patrimonio_map = [
        (31, "  Capital emitido", cap_emitido),
        (32, "  Superávit de capital", superavit),
        (33, "  Utilidad acumulada", util_acum),
        (34, "  Utilidad del periodo", util_periodo),
        (35, "  Reservas", reservas),
    ]

    for r, lbl, val in patrimonio_map:
        ws[f"I{r}"].value = lbl; ws[f"I{r}"].font = F(size=13); ws[f"I{r}"].alignment = A("left")
        ws[f"K{r}"].value = val
        ws[f"K{r}"].number_format = FMT_COP_XL
        ws[f"K{r}"].font = F(size=13)

    ws["I36"].value = "TOTAL PATRIMONIO"; ws["I36"].font = F(bold=True, size=14)
    ws["K36"].value = total_patrimonio if total_patrimonio != 0 else None
    ws["K36"].number_format = FMT_SUB_XL
    ws["K36"].font = F(bold=True, size=14)
    ws["K36"].border = Border(bottom=thin())

    total_pas_patrimonio = total_pas + total_patrimonio
    ws["I38"].value = "TOTAL PASIVOS Y PATRIMONIO"; ws["I38"].font = F(bold=True, size=14)
    ws["K38"].value = total_pas_patrimonio; ws["K38"].number_format = FMT_COP_XL
    ws["K38"].font = F(bold=True, size=14); ws["K38"].border = Border(bottom=double())

    ws["B40"].value = "Las notas adjuntas forman parte integral de estos estados financieros."
    ws["B40"].font = F(size=11)


def generar_hoja_eri(ws, empresa, nit, periodo, totales):
    anchos = {"A":0.9,"B":43.9,"C":6.1,"D":16.6,"E":2.6,"F":16.6,"G":9.0}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w
    for r in range(1, 35):
        ws.row_dimensions[r].height = 15.0
    for r, h in [(1,26.1),(2,20.1),(3,20.1),(4,20.1),(5,20.1),(6,18.0)]:
        ws.row_dimensions[r].height = h

    for r, txt, bold in [(1,empresa,True),(2,nit,False),
                         (3,"ESTADO DE RESULTADOS INTEGRAL",True),
                         (4,periodo,False),(5,"(En pesos colombianos - $)",False)]:
        c = ws[f"B{r}"]; c.value = txt; c.font = F(bold=bold, size=18)

    for col, val in [("D","Acumulado"),("F","Acumulado")]:
        ws[f"{col}6"].value = val; ws[f"{col}6"].font = F(bold=True, size=14)
        ws[f"{col}6"].alignment = A("center")

    ws["C7"].value = "NOTA"; ws["C7"].font = F(bold=True, size=13); ws["C7"].alignment = A("center")
    for col, val in [("D",2025),("F",2024)]:
        ws[f"{col}7"].value = val; ws[f"{col}7"].font = F(bold=True, size=13)
        ws[f"{col}7"].alignment = A("center")
    for col in ["D","F"]:
        ws[f"{col}8"].value = "$"; ws[f"{col}8"].font = F(bold=True, size=13)
        ws[f"{col}8"].alignment = A("center")

    # Valores con signo contable
    ing_ord    = abs(totales.get("INGRESOS DE ACTIVIDADES ORDINARIAS", 0))
    costo_vtas = abs(totales.get("COSTO DE VENTAS", 0))
    otros_ing  = abs(totales.get("OTROS INGRESOS", 0))
    gtos_adm   = abs(totales.get("GASTOS DE ADMINISTRACION", 0))
    gtos_venta = abs(totales.get("GASTOS DE VENTA", 0))
    otros_gto  = abs(totales.get("OTROS GASTOS", 0))
    ing_fin    = abs(totales.get("INGRESOS FINANCIEROS", 0))
    gto_fin    = abs(totales.get("GASTOS FINANCIEROS", 0))
    # Diferencia en cambio neta: puede ser ganancia o pérdida según el signo
    # natural del saldo (crédito = ganancia, débito = pérdida).
    dif_cambio = -totales.get("DIFERENCIA EN CAMBIO NETA", 0)
    provision  = abs(totales.get("PROVISION DE IMPUESTOS", 0))

    ganancia  = ing_ord - costo_vtas   # Ganancia bruta = Ingresos - Costo de ventas (antes ignoraba el costo)
    util_ai   = (ganancia + otros_ing - gtos_adm - gtos_venta - otros_gto
                 + ing_fin - gto_fin + dif_cambio)
    util_per  = util_ai - provision

    lineas = [
        (10, "Ingresos de actividades ordinarias", 13,   ing_ord,    False, None),
        (11, "Costo de ventas",                    None, -costo_vtas,False, None),
        (12, "Ganancia bruta",                      None, ganancia,  True,  None),
        (13, "Gastos de venta",                     None, -gtos_venta,False, None),
        (14, "Otros ingresos",                      14,  otros_ing,  False, None),
        (15, "Gastos de administración",            15, -gtos_adm,   False, None),
        (16, "Otros gastos",                        16, -otros_gto,  False, None),
        (17, "Ingresos financieros",                17,  ing_fin,    False, None),
        (18, "Gastos financieros",                  16, -gto_fin,    False, None),
        (19, "Diferencia en cambio neta",           None, dif_cambio,False, None),
        (21, "Utilidad antes de impuesto",          None, util_ai,   True,  None),
        (23, "Ingreso (gasto) por impuesto",        19, -provision,  False, None),
        (25, "Utilidad (pérdida) del periodo",      None, util_per,  True,  None),
        (29, "Resultado integral total",            None, util_per,  True,  "double"),
    ]

    for row_n, label, nota, val, bold, border in lineas:
        ws[f"B{row_n}"].value = label; ws[f"B{row_n}"].font = F(bold=bold, size=13)
        if nota:
            ws[f"C{row_n}"].value = nota; ws[f"C{row_n}"].font = F(size=13)
            ws[f"C{row_n}"].alignment = A("center")
        c = ws[f"D{row_n}"]
        # Se muestra siempre el valor (incluido 0) para que el concepto aparezca
        # en la plantilla; FMT_COP_XL despliega "-" cuando el saldo es cero.
        c.value = val
        c.number_format = FMT_COP_XL; c.font = F(bold=bold, size=13)
        if border == "double": c.border = Border(bottom=double())

    ws["B31"].value = "Las notas adjuntas forman parte integral de estos estados financieros."
    ws["B31"].font = F(size=12)


def generar_hoja_anexo(ws, title, empresa, df):
    """Genera hojas de Anexo ESF y ERI con diseño profesional, autofiltros y anchos amplios."""
    c_header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    c_subhead_fill = PatternFill(start_color="2E6DA4", end_color="2E6DA4", fill_type="solid")
    c_alt_fill     = PatternFill(start_color="F4F8FA", end_color="F4F8FA", fill_type="solid")
    c_total_fill   = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    font_title  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data   = Font(name="Calibri", size=11, color="1E3A5F")
    font_total  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    border_thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )
    border_total = Border(
        top=Side(style="thin", color="1E3A5F"),
        bottom=Side(style="double", color="FFFFFF")
    )

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left", vertical="center")
    align_right  = Alignment(horizontal="right", vertical="center")

    num_cols = len(df.columns) + 1  # 1 para Col A (Grupo)
    last_col_letter = get_column_letter(num_cols)

    # Fila 1: Banner de Título
    ws.row_dimensions[1].height = 28.0
    ws["A1"].value = f"{empresa.upper()} - {title.upper()}"
    ws["A1"].font  = font_title
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    for c in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = c_header_fill

    # Fila 2: Separador
    ws.row_dimensions[2].height = 10.0

    # Fila 4: Cabecera de tabla con Autofiltros
    ws.row_dimensions[4].height = 26.0
    headers = ["Grupo"] + list(df.columns)
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx)
        c.value = h
        c.font  = font_header
        c.fill  = c_subhead_fill
        c.alignment = align_center

    # Filas de datos desde fila 5
    for r_idx, (idx_val, row) in enumerate(df.iterrows(), start=5):
        ws.row_dimensions[r_idx].height = 20.0
        is_total_row = str(idx_val) in ["Total general", "Total Activo", "Total Pasivo"]

        # Col A: Grupo
        cA = ws.cell(row=r_idx, column=1)
        cA.value = str(idx_val)
        cA.alignment = align_left

        if is_total_row:
            cA.font   = font_total
            cA.fill   = c_total_fill
            cA.border = border_total
        else:
            cA.font   = font_data
            if r_idx % 2 == 0:
                cA.fill = c_alt_fill
            cA.border = border_thin

        # Columnas de valores
        for c_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=r_idx, column=c_idx)
            is_num = isinstance(val, (int, float)) and not pd.isna(val)
            cell.value = float(val) if is_num else (val if not pd.isna(val) else None)
            cell.number_format = FMT_COP_XL
            cell.alignment = align_right

            if is_total_row:
                cell.font   = font_total
                cell.fill   = c_total_fill
                cell.border = border_total
            else:
                cell.font   = font_data
                if r_idx % 2 == 0:
                    cell.fill = c_alt_fill
                cell.border = border_thin

    # Habilitar Autofiltro en la fila 4
    max_row = 4 + len(df)
    ws.auto_filter.ref = f"A4:{last_col_letter}{max_row}"

    # Anchos de columna amplios para legibilidad perfecta de números sin que se corten
    ws.column_dimensions["A"].width = 68.0
    for col_idx in range(2, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22.0


def _escribir_df_en_hoja(ws, df, index=False):
    """Escribe un DataFrame de Detalle con formato, autofiltros y anchos ajustados."""
    c_subhead_fill = PatternFill(start_color="2E6DA4", end_color="2E6DA4", fill_type="solid")
    c_alt_fill     = PatternFill(start_color="F4F8FA", end_color="F4F8FA", fill_type="solid")

    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data   = Font(name="Calibri", size=10, color="1E3A5F")

    border_thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left   = Alignment(horizontal="left", vertical="center")
    align_right  = Alignment(horizontal="right", vertical="center")

    cols = list(df.columns)
    ws.row_dimensions[1].height = 25.0

    # Header
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.value = str(col_name)
        cell.font  = font_header
        cell.fill  = c_subhead_fill
        cell.alignment = align_center

    # Data
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws.row_dimensions[r_idx].height = 19.0
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            col_name = cols[c_idx - 1]
            if isinstance(val, (int, float)) and not pd.isna(val):
                cell.value = float(val)
                if "Saldo" in col_name or "Monto" in col_name or "Valor" in col_name:
                    cell.number_format = FMT_COP_XL
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
            else:
                cell.value = str(val) if not pd.isna(val) else None
                cell.alignment = align_left

            cell.font = font_data
            if r_idx % 2 == 0:
                cell.fill = c_alt_fill
            cell.border = border_thin

    # AutoFilter
    max_col_letter = get_column_letter(len(cols))
    max_row = 1 + len(df)
    ws.auto_filter.ref = f"A1:{max_col_letter}{max_row}"

    # Auto Column Widths
    for c_idx, col_name in enumerate(cols, 1):
        col_letter = get_column_letter(c_idx)
        max_len = max(len(str(col_name)), int(df[col_name].astype(str).str.len().max()) if not df.empty else 10)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 55)


def generar_excel_eeff(empresa, nit, periodo, saldos_esf, saldos_patrimonio, totales_eri,
                       pivot_eri, df_eri_raw, pivot_esf, df_esf_raw):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Hojas formateadas principales
    generar_hoja_esf(wb.create_sheet("ESF"), empresa, nit, periodo, saldos_esf, saldos_patrimonio)
    generar_hoja_eri(wb.create_sheet("ERI"), empresa, nit, periodo, totales_eri)

    # Hojas de Anexo (pivot mensual) con formato estilizado, autofiltros y anchos amplios
    ws_aesf = wb.create_sheet("Anexo ESF")
    generar_hoja_anexo(ws_aesf, "ANEXO AL ESTADO DE LA SITUACIÓN FINANCIERA", empresa, pivot_esf)

    ws_aeri = wb.create_sheet("Anexo ERI")
    generar_hoja_anexo(ws_aeri, "ANEXO AL ESTADO DE RESULTADOS INTEGRAL", empresa, pivot_eri)

    # Hojas de detalle con formato y autofiltros
    ws_desf = wb.create_sheet("Detalle ESF")
    _escribir_df_en_hoja(ws_desf, df_esf_raw, index=False)

    ws_deri = wb.create_sheet("Detalle ERI")
    _escribir_df_en_hoja(ws_deri, df_eri_raw, index=False)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="header-block">
    <h1>📊 Anexos EEFF – ERI y ESF</h1>
    <p>Genera los estados financieros formateados (ESF y ERI) desde la hoja <strong>terceros_</strong></p>
</div>""", unsafe_allow_html=True)

st.markdown('<div class="upload-section">', unsafe_allow_html=True)
st.markdown('<p class="section-title">📁 Archivo Excel de Anexos EEFF</p>', unsafe_allow_html=True)
st.caption("Debe contener la hoja **terceros_** con columnas: Codigo, Nombre cuenta, Nit, Nombre tercero, Mes, Saldo Mes, Cuenta, Grupo")
uploaded = st.file_uploader("Sube el archivo Excel", type=["xlsx"], label_visibility="collapsed")
st.markdown('</div>', unsafe_allow_html=True)

if not uploaded:
    st.markdown("""<div style="text-align:center;padding:60px 20px;color:#999;">
        <div style="font-size:3rem;margin-bottom:16px;">📂</div>
        <p style="font-size:1.1rem;font-weight:600;">Sube el archivo Excel para generar los estados financieros</p>
    </div>""", unsafe_allow_html=True)
    st.stop()

file_bytes = uploaded.read()
df_eri_raw, pivot_eri, df_esf_raw, pivot_esf, saldos_esf, saldos_patrimonio, totales_eri, ultimo_mes = \
    procesar_archivo(file_bytes)

meses_disp     = [m for m in MESES_ORDEN if m in pivot_eri.columns]
meses_disp_esf = [m for m in MESES_ORDEN if m in pivot_esf.columns]

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""<div style="background:linear-gradient(135deg,#1E3A5F,#2E6DA4);padding:14px 18px;
                border-radius:10px;margin-bottom:16px;color:white;text-align:center;">
        <div style="font-size:1.1rem;font-weight:700;">🏢 Datos de la empresa</div></div>""",
        unsafe_allow_html=True)
    empresa = st.text_input("Nombre empresa", value="MI EMPRESA S.A.S")
    nit     = st.text_input("NIT", value="NIT 000.000.000-0")
    periodo = st.text_input("Período", value="AL 31 DE DICIEMBRE DE 2025 Y 31 DE DICIEMBRE DE 2024")

    st.divider()
    st.markdown("**⚙️ Filtros**")
    todos_meses = st.toggle("Todos los meses", value=True, key="tog_meses")
    if todos_meses:
        meses_sel = meses_disp
    else:
        meses_sel = []
        filas = [meses_disp[i:i+4] for i in range(0, len(meses_disp), 4)]
        for fila in filas:
            cols_sb = st.columns(len(fila))
            for col_sb, mes in zip(cols_sb, fila):
                if col_sb.checkbox(MESES_ABREV[mes], value=True, key=f"m_{mes}"):
                    meses_sel.append(mes)

if not meses_sel:
    st.warning("Selecciona al menos un mes."); st.stop()

# ── Tabs principales ──────────────────────────────────────────────────────────
tab_eri, tab_esf, tab_exportar = st.tabs([
    "📈 Estado de Resultados (ERI)",
    "🏦 Estado de Situación Financiera (ESF)",
    "📄 Exportar EEFF Formateado",
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB ERI
# ══════════════════════════════════════════════════════════════════════════════
with tab_eri:
    cols_eri  = meses_sel + ["Total general"]
    pivot_f   = pivot_eri.loc[
        [g for g in GRUPOS_ERI if g in pivot_eri.index] + ["Total general"], cols_eri]

    ingresos = pivot_f.loc[
        [g for g in ["INGRESOS DE ACTIVIDADES ORDINARIAS","OTROS INGRESOS","INGRESOS FINANCIEROS"]
         if g in pivot_f.index], "Total general"].sum()
    gastos   = pivot_f.loc[
        [g for g in ["COSTO DE VENTAS","GASTOS DE ADMINISTRACION","GASTOS DE VENTA","OTROS GASTOS",
                      "GASTOS FINANCIEROS","PROVISION DE IMPUESTOS"]
         if g in pivot_f.index], "Total general"].sum()
    # "Diferencia en cambio neta" no se suma aquí porque puede ser ganancia o
    # pérdida (signo variable); se ve correctamente en la tabla y en la hoja ERI.
    resultado = ingresos + gastos

    st.markdown("---")
    m1, m2, m3 = st.columns(3)
    with m1:
        st.markdown(f"""<div class="metric-card"><div class="number green">{fmt_cop(abs(ingresos))}</div>
            <div class="label">Total Ingresos</div></div>""", unsafe_allow_html=True)
    with m2:
        st.markdown(f"""<div class="metric-card"><div class="number red">{fmt_cop(gastos)}</div>
            <div class="label">Total Gastos</div></div>""", unsafe_allow_html=True)
    with m3:
        rc = "green" if resultado < 0 else "red"
        rl = "Utilidad" if resultado < 0 else "Pérdida"
        st.markdown(f"""<div class="metric-card"><div class="number {rc}">{fmt_cop(abs(resultado))}</div>
            <div class="label">Resultado · {rl}</div></div>""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="result-block">', unsafe_allow_html=True)
    st.markdown('<p class="section-title">📋 Resumen ERI por Grupo y Mes</p>', unsafe_allow_html=True)
    st.dataframe(pivot_f.style.format(fmt_cop), use_container_width=True,
                 height=min(60+40*len(pivot_f), 500))
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown("---")

    gt1, gt2, gt3 = st.tabs(["📈 Evolución mensual","🥧 Composición","🔍 Detalle por tercero"])
    with gt1:
        fig = go.Figure()
        for grupo in [g for g in GRUPOS_ERI if g in pivot_f.index]:
            vals = [pivot_f.loc[grupo, m] if m in pivot_f.columns else 0 for m in meses_sel]
            fig.add_trace(go.Bar(name=grupo, x=meses_sel, y=[abs(v) for v in vals],
                marker_color=COLORES_ERI.get(grupo,"#95a5a6"),
                customdata=[fmt_cop(v) for v in vals],
                hovertemplate=f"<b>{grupo}</b><br>%{{x}}: %{{customdata}}<extra></extra>"))
        fig.update_layout(barmode="group", title="Evolución mensual por grupo",
            xaxis_title="Mes", yaxis_title="COP", height=420,
            legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1),
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig, use_container_width=True)
    with gt2:
        tot_pie = pivot_f.loc[[g for g in GRUPOS_ERI if g in pivot_f.index],"Total general"]
        fig2 = px.pie(values=tot_pie.abs().values, names=tot_pie.index,
            title="Composición por grupo", color=tot_pie.index,
            color_discrete_map=COLORES_ERI, hole=0.4)
        fig2.update_traces(textposition="outside", textinfo="percent+label")
        fig2.update_layout(height=450, showlegend=False)
        st.plotly_chart(fig2, use_container_width=True)
    with gt3:
        st.markdown('<div class="result-block">', unsafe_allow_html=True)
        gd = st.selectbox("Grupo", GRUPOS_ERI, key="eri_det_g")
        md = st.selectbox("Mes",   ["Todos"]+meses_sel, key="eri_det_m")
        mask = df_eri_raw["Grupo"] == gd
        if md != "Todos": mask &= df_eri_raw["Mes"] == md
        df_d = (df_eri_raw[mask & df_eri_raw["Nombre tercero"].notna()]
            [["Mes","Codigo","Nombre cuenta","Nit","Nombre tercero","Saldo Mes"]]
            .sort_values("Saldo Mes", key=abs, ascending=False).reset_index(drop=True))
        df_d["Saldo Mes"] = df_d["Saldo Mes"].apply(fmt_cop)
        st.dataframe(df_d, use_container_width=True, height=400)
        st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB ESF
# ══════════════════════════════════════════════════════════════════════════════
with tab_esf:
    act_r = [g for g in GRUPOS_ESF_ACTIVO if g in pivot_esf.index]
    pas_r = [g for g in GRUPOS_ESF_PASIVO if g in pivot_esf.index]
    meses_esf_f = [m for m in meses_disp_esf if m in meses_sel] or meses_disp_esf
    cols_esf_f  = meses_esf_f + ["Total general"]

    total_act_v = pivot_esf.loc["Total Activo","Total general"] if "Total Activo" in pivot_esf.index else 0
    total_pas_v = pivot_esf.loc["Total Pasivo","Total general"] if "Total Pasivo" in pivot_esf.index else 0
    saldo_act   = sum(abs(saldos_esf.get(g,0)) for g in GRUPOS_ESF_ACTIVO)
    saldo_pas   = sum(abs(saldos_esf.get(g,0)) for g in GRUPOS_ESF_PASIVO)

    st.markdown("---")
    mc1, mc2, mc3 = st.columns(3)
    with mc1:
        st.markdown(f"""<div class="metric-card"><div class="number green">{fmt_cop(abs(total_act_v))}</div>
            <div class="label">Total Activo (acumulado)</div></div>""", unsafe_allow_html=True)
    with mc2:
        st.markdown(f"""<div class="metric-card"><div class="number red">{fmt_cop(abs(total_pas_v))}</div>
            <div class="label">Total Pasivo (acumulado)</div></div>""", unsafe_allow_html=True)
    with mc3:
        saldo_pat = saldos_patrimonio.get("total_patrimonio", 0.0)
        pc = "purple" if saldo_pat >= 0 else "red"
        st.markdown(f"""<div class="metric-card"><div class="number {pc}">{fmt_cop(saldo_pat)}</div>
            <div class="label">Total Patrimonio ({ultimo_mes})</div></div>""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="result-block">', unsafe_allow_html=True)
    st.markdown('<p class="section-title">📋 ESF por Grupo y Mes</p>', unsafe_allow_html=True)
    frames = []
    if act_r: frames.append(pivot_esf.loc[act_r, cols_esf_f])
    if "Total Activo" in pivot_esf.index: frames.append(pivot_esf.loc[["Total Activo"], cols_esf_f])
    if pas_r: frames.append(pivot_esf.loc[pas_r, cols_esf_f])
    if "Total Pasivo" in pivot_esf.index: frames.append(pivot_esf.loc[["Total Pasivo"], cols_esf_f])
    pivot_esf_f = pd.concat(frames) if frames else pd.DataFrame()

    def style_esf(row):
        if row.name in ("Total Activo","Total Pasivo"):
            return ["font-weight:bold;background-color:#EBF5FB;color:#1E3A5F"]*len(row)
        if row.name in GRUPOS_ESF_ACTIVO: return ["color:#1A9E5C"]*len(row)
        return ["color:#D63B3B"]*len(row)

    st.dataframe(pivot_esf_f.style.format(fmt_cop).apply(style_esf, axis=1),
                 use_container_width=True, height=min(60+40*len(pivot_esf_f), 600))
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown("---")

    et1, et2, et3 = st.tabs(["📊 Activo vs Pasivo mensual","🥧 Composición ESF","🔍 Detalle por tercero"])
    with et1:
        a_v = [pivot_esf.loc[act_r, m].sum() if act_r and m in pivot_esf.columns else 0 for m in meses_disp_esf]
        p_v = [pivot_esf.loc[pas_r, m].sum() if pas_r and m in pivot_esf.columns else 0 for m in meses_disp_esf]
        fig_e1 = go.Figure()
        fig_e1.add_trace(go.Scatter(x=meses_disp_esf, y=a_v, mode="lines+markers", name="Total Activo",
            line=dict(color="#1A9E5C",width=3), customdata=[fmt_cop(v) for v in a_v],
            hovertemplate="<b>Activo</b><br>%{x}: %{customdata}<extra></extra>"))
        fig_e1.add_trace(go.Scatter(x=meses_disp_esf, y=p_v, mode="lines+markers", name="Total Pasivo",
            line=dict(color="#D63B3B",width=3), customdata=[fmt_cop(v) for v in p_v],
            hovertemplate="<b>Pasivo</b><br>%{x}: %{customdata}<extra></extra>"))
        fig_e1.update_layout(title="Evolución mensual Activo vs Pasivo",
            xaxis_title="Mes", yaxis_title="COP", height=420,
            legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1),
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig_e1, use_container_width=True)
    with et2:
        grupos_pie = [g for g in GRUPOS_ESF_ORDEN if g in pivot_esf_f.index]
        tot_pie_e = pivot_esf_f.loc[grupos_pie,"Total general"].abs()
        fig_e2 = go.Figure(go.Pie(
            labels=[GRUPOS_LABEL_ESF.get(g,g) for g in tot_pie_e.index],
            values=tot_pie_e.values,
            marker_colors=[COLORES_ESF.get(g,"#aaa") for g in tot_pie_e.index],
            hole=0.4, textposition="outside", textinfo="percent+label"))
        fig_e2.update_layout(title="Composición ESF", height=500, showlegend=False)
        st.plotly_chart(fig_e2, use_container_width=True)
    with et3:
        st.markdown('<div class="result-block">', unsafe_allow_html=True)
        opc = [g for g in GRUPOS_ESF_ORDEN if g in df_esf_raw["Grupo"].unique()]
        gde = st.selectbox("Grupo ESF", opc, key="esf_det_g")
        mde = st.selectbox("Mes",   ["Todos"]+meses_disp_esf, key="esf_det_m")
        mask_e = df_esf_raw["Grupo"] == gde
        if mde != "Todos": mask_e &= df_esf_raw["Mes"] == mde
        df_de = (df_esf_raw[mask_e & df_esf_raw["Nombre tercero"].notna()]
            [["Mes","Codigo","Nombre cuenta","Nit","Nombre tercero","Saldo Mes"]]
            .sort_values("Saldo Mes", key=abs, ascending=False).reset_index(drop=True))
        df_de["Saldo Mes"] = df_de["Saldo Mes"].apply(fmt_cop)
        st.dataframe(df_de, use_container_width=True, height=400)
        st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB EXPORTAR EEFF FORMATEADO
# ══════════════════════════════════════════════════════════════════════════════
with tab_exportar:
    st.markdown("---")
    st.markdown('<div class="result-block">', unsafe_allow_html=True)
    st.markdown('<p class="section-title">📄 Generar EEFF con formato profesional</p>', unsafe_allow_html=True)
    st.info(
        "Genera un archivo Excel con las hojas **ESF** y **ERI** formateadas según el "
        "estándar contable colombiano, más las hojas de Anexos y Detalle. "
        "Configura los datos de la empresa en el panel lateral antes de exportar."
    )

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"**Vista previa ESF — saldo de cierre ({ultimo_mes})**")
        prev_esf_list = [
            {"Cuenta": NOMBRE_ESF.get(g,g).strip(), "Saldo": fmt_cop(abs(saldos_esf.get(g,0)))}
            for g in GRUPOS_ESF_ORDEN if g in saldos_esf
        ]
        prev_esf_list.append({"Cuenta": "─── PATRIMONIO ───", "Saldo": ""})
        prev_esf_list.append({"Cuenta": "Capital emitido", "Saldo": fmt_cop(saldos_patrimonio.get("capital_emitido",0))})
        prev_esf_list.append({"Cuenta": "Superávit de capital", "Saldo": fmt_cop(saldos_patrimonio.get("superavit_capital",0))})
        prev_esf_list.append({"Cuenta": "Utilidad acumulada", "Saldo": fmt_cop(saldos_patrimonio.get("utilidad_acumulada",0))})
        prev_esf_list.append({"Cuenta": "Utilidad del periodo", "Saldo": fmt_cop(saldos_patrimonio.get("utilidad_periodo",0))})
        prev_esf_list.append({"Cuenta": "TOTAL PATRIMONIO", "Saldo": fmt_cop(saldos_patrimonio.get("total_patrimonio",0))})

        prev_esf = pd.DataFrame(prev_esf_list)
        st.dataframe(prev_esf, use_container_width=True, hide_index=True, height=370)
        tot_pas_pat = saldo_pas + saldos_patrimonio.get("total_patrimonio", 0)
        st.markdown(
            f"<p><b>Total Activo:</b> {fmt_cop(saldo_act)} &nbsp;|&nbsp; <b>Total Pasivo:</b> {fmt_cop(saldo_pas)} &nbsp;|&nbsp; <b>Pasivo + Patrimonio:</b> {fmt_cop(tot_pas_pat)}</p>",
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown("**Vista previa ERI — acumulado**")
        ing_ord    = abs(totales_eri.get("INGRESOS DE ACTIVIDADES ORDINARIAS",0))
        costo_vtas = abs(totales_eri.get("COSTO DE VENTAS",0))
        otros_ing  = abs(totales_eri.get("OTROS INGRESOS",0))
        gtos_adm   = abs(totales_eri.get("GASTOS DE ADMINISTRACION",0))
        gtos_venta = abs(totales_eri.get("GASTOS DE VENTA",0))
        otros_gto  = abs(totales_eri.get("OTROS GASTOS",0))
        ing_fin    = abs(totales_eri.get("INGRESOS FINANCIEROS",0))
        gto_fin    = abs(totales_eri.get("GASTOS FINANCIEROS",0))
        dif_cambio = -totales_eri.get("DIFERENCIA EN CAMBIO NETA",0)
        provision  = abs(totales_eri.get("PROVISION DE IMPUESTOS",0))
        ganancia_bruta = ing_ord - costo_vtas
        util_ai  = (ganancia_bruta + otros_ing - gtos_adm - gtos_venta - otros_gto
                    + ing_fin - gto_fin + dif_cambio)
        util_per = util_ai - provision
        prev_eri = pd.DataFrame({
            "Línea": [NOMBRE_ERI.get(g,g) for g in GRUPOS_ERI] +
                     ["─────────────────","Utilidad antes de impuesto","─────────────────","Utilidad del periodo"],
            "Valor": [fmt_cop(-totales_eri.get(g,0)) if g=="DIFERENCIA EN CAMBIO NETA"
                      else fmt_cop(abs(totales_eri.get(g,0))) for g in GRUPOS_ERI] +
                     ["",fmt_cop(util_ai),"",fmt_cop(util_per)],
        })
        st.dataframe(prev_eri, use_container_width=True, hide_index=True, height=370)

    st.markdown("---")

    # El resultado se guarda en session_state para que NO desaparezca cuando
    # Streamlit vuelve a ejecutar el script tras presionar "Descargar"
    # (download_button dispara un rerun completo; si el archivo generado
    # solo vive dentro del `if st.button("Generar")`, en ese rerun el botón
    # vuelve a ser False y todo el bloque desaparece — eso es lo que se
    # sentía como "se sale y envía a otro lugar").
    if st.button("⚙️ Generar archivo EEFF formateado", type="primary", use_container_width=True):
        with st.spinner("Generando archivo Excel…"):
            buf_eeff = generar_excel_eeff(
                empresa, nit, periodo,
                saldos_esf, saldos_patrimonio, totales_eri,
                pivot_eri, df_eri_raw, pivot_esf, df_esf_raw,
            )
        st.session_state["eeff_buffer"] = buf_eeff.getvalue()
        st.session_state["eeff_filename"] = f"EEFF_Formateado_{empresa.strip().replace(' ','_')}.xlsx"

    if st.session_state.get("eeff_buffer"):
        st.success("✅ Archivo generado — 6 hojas: ESF, ERI, Anexo ESF, Anexo ERI, Detalle ESF, Detalle ERI")
        st.download_button(
            label="📥 Descargar EEFF_Formateado.xlsx",
            data=st.session_state["eeff_buffer"],
            file_name=st.session_state.get("eeff_filename", "EEFF_Formateado.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="dl_eeff_btn",
        )
    st.markdown('</div>', unsafe_allow_html=True)

