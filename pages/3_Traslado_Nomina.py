import streamlit as st
import pandas as pd
import io
import openpyxl
from copy import copy
from datetime import datetime

if not st.session_state.get("authentication_status"):
    st.warning("⚠️ Debe iniciar sesión primero.")
    st.page_link("Inicio.py", label="Ir al login", icon="🔐")
    st.stop()

st.set_page_config(page_title="Traslado Nómina", page_icon="📋", layout="wide")

# ─────────────────────────────────────────────────────────────────────────────
# IMPORTANTE: sube esta cadena CADA vez que cambies la lógica de cálculo.
# Aparece en el sidebar como "Build:". Si en pantalla NO ves esta misma cadena,
# tu app NO está corriendo este archivo (otro proceso, otra ruta u otro puerto).
# ─────────────────────────────────────────────────────────────────────────────
APP_VERSION = "v5-2026-06-19"

st.markdown("""
<style>
    .stApp { background-color: #F5F7FA; }
    .header-block {
        background: linear-gradient(135deg, #1E3A5F 0%, #2E6DA4 100%);
        border-radius: 12px; padding: 28px 36px; margin-bottom: 28px; color: white;
    }
    .header-block h1 { font-size: 1.9rem; margin: 0; font-weight: 700; }
    .header-block p  { margin: 6px 0 0; opacity: .85; font-size: .95rem; }
    .metric-card {
        background: white; border-radius: 10px; padding: 20px 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,.08); text-align: center;
    }
    .metric-card .number { font-size: 2.2rem; font-weight: 700; }
    .metric-card .label  { font-size: .82rem; color: #666; text-transform: uppercase; }
    .green  { color: #1A9E5C; } .red    { color: #D63B3B; }
    .blue   { color: #2E6DA4; } .orange { color: #E07B20; }
    .upload-section {
        background: white; border-radius: 10px; padding: 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,.08); margin-bottom: 20px;
    }
    .section-title {
        font-size: 1rem; font-weight: 600; color: #1E3A5F;
        margin-bottom: 12px; padding-bottom: 8px; border-bottom: 2px solid #E8EDF3;
    }
    .section-title-sn {
        font-size: 1rem; font-weight: 600; color: #2E6DA4;
        margin-bottom: 12px; padding-bottom: 8px; border-bottom: 2px solid #BBDEFB;
    }
    .bloque-box {
        background: white; border-radius: 10px; padding: 20px 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,.08); margin-bottom: 16px;
        border-left: 4px solid #2E6DA4;
    }
    .bloque-box-sn {
        background: white; border-radius: 10px; padding: 20px 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,.08); margin-bottom: 16px;
        border-left: 4px solid #2E6DA4;
    }
    .warn-box {
        background: #FFF8E1; border-left: 4px solid #E07B20;
        border-radius: 8px; padding: 12px 16px; margin: 10px 0; font-size: .88rem;
    }
    .info-box {
        background: #E8F4FD; border-left: 4px solid #2E6DA4;
        border-radius: 8px; padding: 12px 16px; margin: 10px 0; font-size: .88rem;
    }
    #MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR: build-stamp + reset manual (te dice qué código corre y limpia todo)
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.caption(f"🏷️ Build: {APP_VERSION}")
    if st.button("🧹 Limpiar caché y memoria"):
        st.cache_data.clear()
        for k in list(st.session_state.keys()):
            if k != "authentication_status":
                st.session_state.pop(k, None)
        st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# FÓRMULAS DEL DESTINO (solo se usan en modo "formulas")
# ─────────────────────────────────────────────────────────────────────────────
_HOM = "'HOMOLOGACION CONCEPTOS'"

def formulas_fila(r):
    return {
        1:  f"=+VLOOKUP(D{r},{_HOM}!$B$1:$I$81,8,FALSE)",
        9:  f'=+IF(MID(B{r},1,1)="3",-H{r},H{r})',
        10: f"=+VLOOKUP(E{r},CC!$B:$C,2,FALSE)",
        11: f"=+J{r}",
        12: f'=+IF(MID(J{r},1,2)="12",VLOOKUP(D{r},{_HOM}!$B$1:$E$1037,4,0),'
            f'VLOOKUP(D{r},{_HOM}!$B$1:$E$1037,3,0))',
        13: f'=+IF(VLOOKUP(D{r},{_HOM}!$B$1:$G$1148,5,0)="D",'
            f'IF(H{r}<0,"C",VLOOKUP(D{r},{_HOM}!$B$1:$G$1148,5,0)),"C")',
        14: f'=+IF(M{r}="D",ABS(I{r}),0)',
        15: f'=+IF(M{r}="C",ABS(I{r}),0)',
        16: f"=+VLOOKUP(D{r},{_HOM}!$B$1:$G$1148,6,0)",
        17: f'=+IF(M{r}="C","D","C")',
        18: f"=+O{r}",
        19: f"=+N{r}",
        20: f'=+IF(A{r}="PAGO_SS",VLOOKUP(U{r},\'TERCEROS PILA\'!$A$1:$E$35,3,0),E{r})',
    }


# ─────────────────────────────────────────────────────────────────────────────
# LECTURA
# ─────────────────────────────────────────────────────────────────────────────
def detectar_fila_header(ws):
    pistas = ("valor", "nombre", "concepto", "horas", "cod", "cedul",
              "grupo", "id", "fecha")
    for row in ws.iter_rows(min_row=1, max_row=25):
        vals = [str(c.value).lower() if c.value else "" for c in row]
        hits = sum(any(p in v for p in pistas) for v in vals if v)
        if hits >= 2:
            return row[0].row
    return 1


def _parse_hoja(raw, hint_hoja, exacta=False):
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    if exacta:
        hoja = next((s for s in wb.sheetnames if hint_hoja.lower() in s.lower()), None)
        if hoja is None:
            return None, None, None
    else:
        hoja = next(
            (s for s in wb.sheetnames if hint_hoja.lower() in s.lower()),
            wb.sheetnames[0]
        )
    ws = wb[hoja]
    hr = detectar_fila_header(ws)

    headers, seen = [], {}
    for c in range(1, ws.max_column + 1):
        v    = ws.cell(hr, c).value
        name = str(v).strip() if v is not None else f"__COL{c}__"
        if name in seen:
            seen[name] += 1
            name = f"{name}__{seen[name]}"
        else:
            seen[name] = 1
        headers.append(name)

    rows = []
    for r in range(hr + 1, ws.max_row + 1):
        fila = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in fila):
            rows.append(fila)

    df = pd.DataFrame(rows, columns=headers)
    return df, hoja, hr


@st.cache_data(show_spinner="Leyendo archivos...")
def cargar_archivos(raw33, raw32, _ver):
    # _ver = APP_VERSION: al cambiar la versión, la firma del cache cambia y se
    # vuelve a leer (no quedan datos de una versión anterior).
    df_datos,   hoja_datos,   hr_datos   = _parse_hoja(raw33, "dato")
    df_resumen, hoja_resumen, hr_resumen = _parse_hoja(raw33, "resumen", exacta=True)
    df_32,      hoja_32,      hr_32       = _parse_hoja(raw32, "nomina")
    return (df_datos, hoja_datos, hr_datos,
            df_resumen, hoja_resumen, hr_resumen,
            df_32, hoja_32, hr_32)


def col_exacta(df, nombre):
    for col in df.columns:
        if col.strip().upper() == nombre.strip().upper():
            return col
    return None


def col_parcial(df, *palabras):
    for col in df.columns:
        cn = col.strip().upper()
        if all(p.upper() in cn for p in palabras):
            return col
    return None


def automap_33_datos(df):
    return {
        "codigo":   (col_exacta(df, "CODIGOCONCEPTO") or
                     col_parcial(df, "CODIGO", "CONCEPTO") or
                     col_parcial(df, "COD")),
        "concepto": (col_exacta(df, "CONCEPTO") or col_parcial(df, "CONCEPTO")),
        "id":       (col_exacta(df, "ID") or col_parcial(df, "CEDULA") or
                     col_parcial(df, "IDENTIF")),
        "nombre":   (col_exacta(df, "NOMBRECOMPLETO") or
                     col_parcial(df, "NOMBRE", "COMPLETO") or
                     col_parcial(df, "NOMBRE")),
        "horas":    (col_exacta(df, "HORAS") or col_parcial(df, "HORA")),
        "valor":    (col_exacta(df, "VALOR") or col_parcial(df, "VALOR")),
    }


def automap_33_resumen(df):
    cols = list(df.columns)
    return {
        "id":     (col_exacta(df, "ID") or col_parcial(df, "CEDULA") or
                   col_parcial(df, "IDENTIF") or (cols[0] if cols else None)),
        "nombre": (col_exacta(df, "NOMBRE COMPLETO") or
                   col_exacta(df, "NOMBRECOMPLETO") or
                   col_parcial(df, "NOMBRE", "COMPLETO") or
                   col_parcial(df, "NOMBRE") or (cols[1] if len(cols) > 1 else None)),
        "valor":  (col_exacta(df, "VALOR") or col_parcial(df, "VALOR") or
                   (cols[7] if len(cols) > 7 else None)),
    }


def automap_32(df):
    return {
        "cod_dina":   (col_exacta(df, "COD_DINA") or col_parcial(df, "COD_DINA") or
                       col_parcial(df, "COD")),
        "concepto_b": (col_exacta(df, "NOMBRE DEL CONCEPTO") or
                       col_parcial(df, "NOMBRE", "CONCEPTO") or
                       col_parcial(df, "CONCEPTO")),
        "concepto_d": None,
        "cod_helisa": (col_exacta(df, "COD_HELISA") or col_parcial(df, "COD_HELISA") or
                       col_parcial(df, "HELISA")),
        "cedula":     (col_exacta(df, "CEDULA") or col_parcial(df, "CEDULA") or
                       col_parcial(df, "IDENTIF")),
        "nombre_emp": (col_exacta(df, "NOMBRE EMPLEADO") or
                       col_parcial(df, "NOMBRE", "EMPLEADO") or
                       col_parcial(df, "NOMBRE")),
        "horas":      (col_exacta(df, "HORAS") or col_parcial(df, "HORA")),
        "valor":      (col_exacta(df, "VALOR") or col_parcial(df, "VALOR")),
    }


def build_col_idx(ws, hr):
    col_idx, seen = {}, {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hr, c).value
        if v:
            name = str(v).strip()
            if name in seen:
                seen[name] += 1
                col_idx[f"{name}__{seen[name]}"] = c
            else:
                seen[name] = 1
                col_idx[name] = c
    return col_idx


# ─────────────────────────────────────────────────────────────────────────────
# CÁLCULO EN PYTHON (modo "valores")
# ─────────────────────────────────────────────────────────────────────────────
COLS_CALC = [1, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]


def _norm_key(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s.upper()


def construir_lookups(raw_32):
    wb = openpyxl.load_workbook(io.BytesIO(raw_32), data_only=True)
    H = wb["HOMOLOGACION CONCEPTOS"]
    hom = {}
    for r in range(1, H.max_row + 1):
        k = _norm_key(H.cell(r, 2).value)
        if k is None or k in hom:
            continue
        hom[k] = {
            "cuenta":        H.cell(r, 4).value,
            "doce":          H.cell(r, 5).value,
            "nat":           H.cell(r, 6).value,
            "contrapartida": H.cell(r, 7).value,
            "fuente":        H.cell(r, 9).value,
        }
    C = wb["CC"]
    cc = {}
    for r in range(1, C.max_row + 1):
        k = _norm_key(C.cell(r, 2).value)
        if k is not None and k not in cc:
            cc[k] = C.cell(r, 3).value
    T = wb["TERCEROS PILA"]
    ter = {}
    for r in range(1, T.max_row + 1):
        k = _norm_key(T.cell(r, 1).value)
        if k is not None and k not in ter:
            ter[k] = T.cell(r, 3).value
    return hom, cc, ter


def calcular_fila(B, D, E, H, U, hom, cc, ter):
    h = H if isinstance(H, (int, float)) else 0
    neto = -h if (B is not None and str(B).strip()[:1] == "3") else h
    rec    = hom.get(_norm_key(D))
    fuente = rec["fuente"] if rec else None
    j      = cc.get(_norm_key(E))
    if rec is None:
        nat = None
    elif rec["nat"] == "D":
        nat = "C" if h < 0 else "D"
    else:
        nat = "C"
    if j is not None and str(j)[:2] == "12":
        puc = rec["doce"] if rec else None
    else:
        puc = rec["cuenta"] if rec else None
    n = abs(neto) if nat == "D" else 0
    o = abs(neto) if nat == "C" else 0
    contrap = rec["contrapartida"] if rec else None
    q = "D" if nat == "C" else "C"
    t = (ter.get(_norm_key(U)) if fuente == "PAGO_SS" else E)
    return {1: fuente, 9: neto, 10: j, 11: j, 12: puc, 13: nat,
            14: n, 15: o, 16: contrap, 17: q, 18: o, 19: n, 20: t}


# ─────────────────────────────────────────────────────────────────────────────
# TRASLADO
# ─────────────────────────────────────────────────────────────────────────────
def ejecutar_traslado_doble(
    raw_32, hoja_32, hr_32,
    df_datos, mapa_src_datos, mapa_dst_datos,
    df_resumen, mapa_src_resumen, mapa_dst_resumen,
    col_deteccion_sn,
    modo="valores",
):
    wb = openpyxl.load_workbook(io.BytesIO(raw_32))
    ws = wb[hoja_32]
    col_idx = build_col_idx(ws, hr_32)
    warns   = []

    # BLOQUE 1: hoja DATOS, TODAS las filas (sin tope)
    primera_dato = hr_32 + 1
    df_d = df_datos.reset_index(drop=True)
    n_ok = 0
    for i, row in df_d.iterrows():
        fila_excel = primera_dato + i
        for clave, src_col in mapa_src_datos.items():
            dst_col = mapa_dst_datos.get(clave)
            if not src_col or not dst_col:
                continue
            cn = col_idx.get(dst_col)
            if cn is None:
                msg = f"Col destino '{dst_col}' no encontrada"
                if msg not in warns:
                    warns.append(msg)
                continue
            val = row.get(src_col)
            if isinstance(val, float) and pd.isna(val):
                val = None
            ws.cell(fila_excel, cn, value=val)
        n_ok += 1
    ultima_dato = primera_dato + len(df_d) - 1

    # BLOQUE 2: filas SN
    n_sn_escritas = 0
    if df_resumen is not None and len(df_resumen) > 0:
        cn_dina = col_idx.get(col_deteccion_sn)
        filas_sn = []
        if cn_dina:
            for r in range(hr_32 + 1, ws.max_row + 1):
                val_dina = ws.cell(r, cn_dina).value
                if val_dina is not None and str(val_dina).strip().upper() == "SN":
                    filas_sn.append(r)
        filas_validas = []
        for _, row in df_resumen.iterrows():
            row_str = " ".join(str(v).upper() for v in row.values)
            if "NOMBRECOMPLETO" in row_str or "CEDULA" in row_str or "SALARIO NETO" in row_str:
                continue
            filas_validas.append(row)
        df_resumen_limpio = (pd.DataFrame(filas_validas)
                             if filas_validas
                             else pd.DataFrame(columns=df_resumen.columns))
        n_sn = min(len(df_resumen_limpio), len(filas_sn))
        for i in range(n_sn):
            fila_excel = filas_sn[i]
            row = df_resumen_limpio.iloc[i]
            for clave, src_col in mapa_src_resumen.items():
                dst_col = mapa_dst_resumen.get(clave)
                if not src_col or not dst_col:
                    continue
                cn = col_idx.get(dst_col)
                if cn is None:
                    continue
                val = row.get(src_col)
                if isinstance(val, float) and pd.isna(val):
                    val = None
                ws.cell(fila_excel, cn, value=val)
            n_sn_escritas += 1
        if len(df_resumen_limpio) > len(filas_sn):
            warns.append(
                f"Nomina resumen tiene {len(df_resumen_limpio)} filas válidas "
                f"pero solo hay {len(filas_sn)} filas SN en el 3.2."
            )

    # COLUMNAS CALCULADAS
    if modo in ("valores", "formulas") and ultima_dato >= primera_dato:
        fila_ref = primera_dato
        for fe in range(primera_dato + 1, ultima_dato + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(fe, c)._style = copy(ws.cell(fila_ref, c)._style)

        if modo == "formulas":
            for fe in range(primera_dato, ultima_dato + 1):
                for c, f in formulas_fila(fe).items():
                    ws.cell(fe, c, value=f)
        else:  # valores
            hom, cc, ter = construir_lookups(raw_32)
            for fe in range(primera_dato, ultima_dato + 1):
                B = ws.cell(fe, 2).value
                D = ws.cell(fe, 4).value
                E = ws.cell(fe, 5).value
                H = ws.cell(fe, 8).value
                U = ws.cell(fe, 21).value
                vals = calcular_fila(B, D, E, H, U, hom, cc, ter)
                for c, v in vals.items():
                    ws.cell(fe, c, value=v)

        # Limpiar residuos de la plantilla por debajo de la última fila escrita
        for fe in range(ultima_dato + 1, ws.max_row + 1):
            for c in COLS_CALC:
                if ws.cell(fe, c).value is not None:
                    ws.cell(fe, c, value=None)

    # ── AUTOCHEQUEO: cuántas filas de datos quedaron sin Vr Débito Y Crédito ──
    sin_calc = 0
    if modo in ("valores", "formulas"):
        for fe in range(primera_dato, ultima_dato + 1):
            if ws.cell(fe, 2).value in (None, ""):
                continue
            n_val = ws.cell(fe, 14).value
            o_val = ws.cell(fe, 15).value
            if n_val in (None, "") and o_val in (None, ""):
                sin_calc += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), n_ok, n_sn_escritas, ultima_dato, warns, sin_calc


# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-block">
    <h1>📋 Traslado de Nómina — 3.3 → 3.2</h1>
    <p><b>Bloque 1:</b> hoja DATOS → todas las filas del 3.2 &nbsp;|&nbsp;
       <b>Bloque 2:</b> hoja Nomina resumen → filas SN &nbsp;|&nbsp;
       <b>Vr Débito/Crédito</b> extendidos a todas las filas</p>
</div>""", unsafe_allow_html=True)

st.markdown("---")

col1, col2 = st.columns(2)
with col1:
    st.markdown('<div class="upload-section"><p class="section-title">📄 Archivo 3.3 — Fuente</p>',
                unsafe_allow_html=True)
    st.caption("Se leerán las hojas **DATOS** y **Nomina resumen**")
    f33 = st.file_uploader("33", type=["xlsx", "xls"], key="u33",
                           label_visibility="collapsed")
    st.markdown('</div>', unsafe_allow_html=True)
with col2:
    st.markdown('<div class="upload-section"><p class="section-title">📊 Archivo 3.2 — Destino (hoja NÓMINA)</p>',
                unsafe_allow_html=True)
    st.caption("Bloque 1: todas las filas · Bloque 2: filas con COD_DINA = SN")
    f32 = st.file_uploader("32", type=["xlsx", "xls"], key="u32",
                           label_visibility="collapsed")
    st.markdown('</div>', unsafe_allow_html=True)

if f33 and f32:
    raw33  = f33.getvalue()
    raw_32 = f32.getvalue()

    (df_datos, hoja_datos, hr_datos,
     df_resumen, hoja_resumen, hr_resumen,
     df_32, hoja_32, hr_32) = cargar_archivos(raw33, raw_32, APP_VERSION)

    m_datos   = automap_33_datos(df_datos)
    m_resumen = automap_33_resumen(df_resumen) if df_resumen is not None else {"id": None, "nombre": None, "valor": None}
    m32       = automap_32(df_32)

    with st.expander("🔎 Estructura detectada"):
        t1, t2, t3 = st.tabs(["3.3 — DATOS", "3.3 — Nomina resumen", "3.2 — NÓMINA"])
        with t1:
            st.caption(f"Hoja: **{hoja_datos}** | {len(df_datos)} filas")
            st.dataframe(df_datos.head(8), use_container_width=True)
        with t2:
            if df_resumen is not None:
                st.caption(f"Hoja: **{hoja_resumen}** | {len(df_resumen)} filas")
                st.dataframe(df_resumen.head(8), use_container_width=True)
            else:
                st.warning("No se encontró la hoja 'Nomina resumen' en el 3.3")
        with t3:
            st.caption(f"Hoja: **{hoja_32}** | {len(df_32)} filas")
            st.dataframe(df_32.head(8), use_container_width=True)

    with st.expander("⚙️ Ajustar mapeo de columnas"):
        op_datos   = list(df_datos.columns)
        op_resumen = list(df_resumen.columns) if df_resumen is not None else ["—"]
        op32       = list(df_32.columns)

        def idx(lst, val):
            return lst.index(val) if val in lst else 0

        st.markdown("**🟢 Bloque 1 — Fuente: 3.3 DATOS**")
        a, b, c = st.columns(3)
        with a:
            m_datos["codigo"]   = st.selectbox("CODIGOCONCEPTO", op_datos, index=idx(op_datos, m_datos["codigo"]),   key="d_cod")
            m_datos["id"]       = st.selectbox("ID",             op_datos, index=idx(op_datos, m_datos["id"]),       key="d_id")
        with b:
            m_datos["concepto"] = st.selectbox("CONCEPTO",       op_datos, index=idx(op_datos, m_datos["concepto"]), key="d_con")
            m_datos["nombre"]   = st.selectbox("NOMBRECOMPLETO", op_datos, index=idx(op_datos, m_datos["nombre"]),   key="d_nom")
        with c:
            m_datos["horas"]    = st.selectbox("HORAS", op_datos, index=idx(op_datos, m_datos["horas"]), key="d_hor")
            m_datos["valor"]    = st.selectbox("VALOR", op_datos, index=idx(op_datos, m_datos["valor"]), key="d_val")

        st.markdown("---")
        st.markdown("**🔵 Bloque 2 — Fuente: 3.3 Nomina resumen**")
        r1, r2, r3 = st.columns(3)
        with r1:
            m_resumen["id"]     = st.selectbox("ID / Cédula",     op_resumen, index=idx(op_resumen, m_resumen.get("id")),     key="r_id")
        with r2:
            m_resumen["nombre"] = st.selectbox("Nombre Completo", op_resumen, index=idx(op_resumen, m_resumen.get("nombre")), key="r_nom")
        with r3:
            m_resumen["valor"]  = st.selectbox("Valor (col H)",   op_resumen, index=idx(op_resumen, m_resumen.get("valor")),  key="r_val")

        st.markdown("---")
        st.markdown("**Destino — 3.2 NÓMINA**")
        d, e, f_ = st.columns(3)
        with d:
            m32["cod_dina"]   = st.selectbox("COD_DINA",   op32, index=idx(op32, m32["cod_dina"]),   key="t_cod")
            m32["cod_helisa"] = st.selectbox("COD_HELISA", op32, index=idx(op32, m32["cod_helisa"]), key="t_hel")
            m32["cedula"]     = st.selectbox("CEDULA",     op32, index=idx(op32, m32["cedula"]),     key="t_ced")
        with e:
            m32["concepto_b"] = st.selectbox("NOMBRE DEL CONCEPTO", op32, index=idx(op32, m32["concepto_b"]), key="t_conb")
        with f_:
            m32["nombre_emp"] = st.selectbox("NOMBRE EMPLEADO", op32, index=idx(op32, m32["nombre_emp"]), key="t_nom")
            m32["horas"]      = st.selectbox("HORAS",           op32, index=idx(op32, m32["horas"]),      key="t_hor")
            m32["valor"]      = st.selectbox("VALOR",           op32, index=idx(op32, m32["valor"]),      key="t_val")

    with st.expander("🧮 Opciones de columnas calculadas", expanded=True):
        usar_formulas = st.checkbox(
            "Usar fórmulas de Excel en lugar de valores "
            "(solo si SIEMPRE abrirás el archivo en Excel)", value=False)
        modo = "formulas" if usar_formulas else "valores"
        if modo == "valores":
            st.markdown('<div class="info-box">✅ <b>Modo Valores</b>: Vr Débito / Vr '
                        'Crédito se calculan en Python y se escriben como números. '
                        'Se ven en cualquier programa, sin recalcular.</div>',
                        unsafe_allow_html=True)
        else:
            st.markdown('<div class="warn-box">⚠️ <b>Modo Fórmulas</b>: las columnas '
                        'se ven <b>vacías (naranja)</b> hasta que Excel recalcule al '
                        'abrir. Usa Valores si dudas.</div>', unsafe_allow_html=True)

    st.markdown("---")
    if st.button("🚀 Ejecutar traslado completo y generar 3.2 actualizado",
                 type="primary", use_container_width=True):

        # ── Limpieza forzada en cada ejecución: nada de resultados viejos ──
        st.cache_data.clear()
        st.session_state.pop("traslado_resultado", None)

        mapa_src_datos = {
            "cod_dina": m_datos["codigo"], "concepto_b": m_datos["concepto"],
            "concepto_d": m_datos["concepto"], "cod_helisa": m_datos["codigo"],
            "cedula": m_datos["id"], "nombre_emp": m_datos["nombre"],
            "horas": m_datos["horas"], "valor": m_datos["valor"],
        }
        mapa_dst_datos = {
            "cod_dina": m32["cod_dina"], "concepto_b": m32["concepto_b"],
            "concepto_d": m32["concepto_d"], "cod_helisa": m32["cod_helisa"],
            "cedula": m32["cedula"], "nombre_emp": m32["nombre_emp"],
            "horas": m32["horas"], "valor": m32["valor"],
        }
        mapa_src_resumen = {"cedula": m_resumen.get("id"),
                            "nombre_emp": m_resumen.get("nombre"),
                            "valor": m_resumen.get("valor")}
        mapa_dst_resumen = {"cedula": m32["cedula"],
                            "nombre_emp": m32["nombre_emp"],
                            "valor": m32["valor"]}

        with st.spinner("Procesando traslado y calculando columnas..."):
            resultado, n_ok, n_sn, ultima, warns, sin_calc = ejecutar_traslado_doble(
                raw_32, hoja_32, hr_32,
                df_datos, mapa_src_datos, mapa_dst_datos,
                df_resumen, mapa_src_resumen, mapa_dst_resumen,
                m32["cod_dina"], modo=modo,
            )

        st.session_state["traslado_resultado"] = {
            "bytes": resultado, "n_ok": n_ok, "n_sn": n_sn, "ultima": ultima,
            "warns": warns, "sin_calc": sin_calc, "hoja": hoja_32, "modo": modo,
            "version": APP_VERSION,
            "nombre": f"3_2_actualizado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        }

    res = st.session_state.get("traslado_resultado")
    if res and (res.get("version") != APP_VERSION or res.get("modo") != modo):
        st.info("ℹ️ Cambió la configuración. Pulsa **🚀 Ejecutar** de nuevo.")
        res = None

    if res:
        if res["warns"]:
            with st.expander(f"⚠️ {len(res['warns'])} advertencias"):
                for w in res["warns"]:
                    st.warning(w)

        # ── AUTOCHEQUEO visible: debe ser 0 ──
        cM, cT = st.columns(2)
        with cM:
            st.metric("Filas SIN Vr Débito/Crédito", res.get("sin_calc", 0),
                      help="Debe ser 0. Si es >0, hay filas cuyo COD_HELISA no "
                           "está en HOMOLOGACION o el VALOR llegó vacío.")
        with cT:
            st.metric("Última fila escrita", res.get("ultima", 0))

        if res.get("sin_calc", 0) == 0:
            st.success(f"✅ {res['n_ok']} filas DATOS · {res['n_sn']} filas SN · "
                       f"Vr Débito/Crédito calculados hasta la fila {res['ultima']} "
                       f"en **{res['hoja']}** (build {res['version']}).")
        else:
            st.warning(f"⚠️ {res['sin_calc']} filas quedaron sin Vr Débito/Crédito. "
                       "Revisa COD_HELISA / VALOR de esas filas.")

        if res["modo"] == "formulas":
            st.info("📌 Modo Fórmulas: ábrelo en Excel para que recalcule (un visor "
                    "que no recalcula las verá naranjas y vacías).")

        # Vista previa de las últimas filas generadas
        if res["modo"] == "valores":
            try:
                wsp = openpyxl.load_workbook(io.BytesIO(res["bytes"]), data_only=True)[res["hoja"]]
                ult = res["ultima"]; ini = max(2, ult - 7)
                filas = [{
                    "Fila": r, "COD_DINA": wsp.cell(r, 2).value,
                    "CONCEPTO": wsp.cell(r, 3).value, "VALOR": wsp.cell(r, 8).value,
                    "Vr DEBITO": wsp.cell(r, 14).value, "Vr CREDITO": wsp.cell(r, 15).value,
                } for r in range(ini, ult + 1)]
                st.caption("🔍 Últimas filas generadas:")
                st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)
            except Exception:
                pass

        st.download_button(
            "⬇️ Descargar 3.2 actualizado", data=res["bytes"],
            file_name=res["nombre"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

else:
    st.session_state.pop("traslado_resultado", None)
    st.markdown("""
    <div style="text-align:center;padding:60px 20px;color:#999;">
        <div style="font-size:3rem;margin-bottom:16px;">📂</div>
        <p style="font-size:1.1rem;font-weight:600;">Carga los dos archivos para iniciar</p>
        <p style="font-size:.9rem;">3.3 → <b>DATOS</b> y <b>Nomina resumen</b> ·
           3.2 → hoja <b>NÓMINA</b></p>
    </div>""", unsafe_allow_html=True)
